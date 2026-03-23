#!/usr/bin/env python3
"""
LU/Parliament - Luxembourg Chamber of Deputies (Chambre des Députés)

Fetches parliamentary written questions and answers with full text from the
Luxembourg Open Data platform (data.public.lu).

Data source: https://data.public.lu/datasets/66163fc900322d4991ed3a9c/
License: CC Zero (CC0)
"""

import argparse
import io
import json
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Generator, Dict, Any, Optional

import requests
from openpyxl import load_workbook

# Configuration
DATASET_ID = "66163fc900322d4991ed3a9c"
BASE_URL = "https://data.public.lu/api/1"
SOURCE_ID = "LU/Parliament"

# XLSX resources for each legislature period (contains full text, unlike TXT which is truncated)
RESOURCES = {
    "2013_2018": "https://download.data.public.lu/resources/ensemble-des-questions-et-reponses-parlementaires-incluant-le-texte-complet/20250428-165929/opendata-qp-ecrite-2013-2018.xlsx",
    "2018_2023": "https://download.data.public.lu/resources/ensemble-des-questions-et-reponses-parlementaires-incluant-le-texte-complet/20250428-165934/opendata-qp-ecrite-2018-2023.xlsx",
    "2023_2028": "https://download.data.public.lu/resources/ensemble-des-questions-et-reponses-parlementaires-incluant-le-texte-complet/20250428-165933/opendata-qp-ecrite-2023-2028.xlsx",
}

# Column mapping (0-indexed)
# Headers: Legislature, NumQuestion, DateDepotQ, AuteurQ, PartiPolitique, TexteQ, TitreQuestion,
#          LangueQ, DateReponse, AuteurR, TexteR, LangueR, lnk_Question, lnk_DocQ, lnk_DocR
COL_LEGISLATURE = 0
COL_QUESTION_NUM = 1
COL_QUESTION_DATE = 2
COL_AUTHOR = 3
COL_PARTY = 4
COL_QUESTION_TEXT = 5
COL_TITLE = 6
COL_LANGUAGE_Q = 7
COL_ANSWER_DATE = 8
COL_ANSWER_AUTHOR = 9
COL_ANSWER_TEXT = 10
COL_LANGUAGE_R = 11
COL_URL = 12
COL_PDF_Q = 13
COL_PDF_R = 14


def fetch_xlsx(url: str, session: requests.Session) -> bytes:
    """Download an XLSX file from data.public.lu."""
    print(f"  Downloading: {url}")
    response = session.get(url, timeout=120)
    response.raise_for_status()
    print(f"  File size: {len(response.content):,} bytes")
    return response.content


def parse_date(date_val) -> Optional[str]:
    """Parse date to ISO 8601 format."""
    if date_val is None:
        return None

    # Handle datetime objects from Excel
    if isinstance(date_val, datetime):
        return date_val.strftime('%Y-%m-%d')

    date_str = str(date_val).strip()
    if not date_str:
        return None

    try:
        # Handle DD/MM/YYYY format
        parts = date_str.split('/')
        if len(parts) == 3:
            day, month, year = parts
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        # Handle YYYY-MM-DD format
        if '-' in date_str and len(date_str) == 10:
            return date_str
    except:
        pass
    return None


def clean_text(text) -> str:
    """Clean text content, handling None and converting pipes to newlines."""
    if text is None:
        return ''
    text = str(text).strip()
    # Replace pipe separators with newlines (used in multilingual content)
    text = text.replace('|', '\n')
    return text


def parse_row(row: tuple) -> Optional[Dict[str, Any]]:
    """Parse an Excel row tuple into a dictionary."""
    if len(row) < 10:
        return None

    # Pad row if needed
    row = list(row)
    while len(row) < 15:
        row.append(None)

    legislature = str(row[COL_LEGISLATURE] or '').strip()
    question_num = str(row[COL_QUESTION_NUM] or '').strip()

    # Skip header row or empty rows
    if not question_num or question_num == 'NumQuestion' or legislature == 'Legislature':
        return None

    return {
        'legislature': legislature,
        'question_number': question_num,
        'question_date': row[COL_QUESTION_DATE],
        'author': str(row[COL_AUTHOR] or '').strip(),
        'political_party': str(row[COL_PARTY] or '').strip(),
        'question_text': clean_text(row[COL_QUESTION_TEXT]),
        'title': str(row[COL_TITLE] or '').strip(),
        'languages': str(row[COL_LANGUAGE_Q] or '').strip(),
        'answer_date': row[COL_ANSWER_DATE],
        'answer_author': str(row[COL_ANSWER_AUTHOR] or '').strip(),
        'answer_text': clean_text(row[COL_ANSWER_TEXT]),
        'answer_language': str(row[COL_LANGUAGE_R] or '').strip(),
        'question_url': str(row[COL_URL] or '').strip(),
        'question_pdf_url': str(row[COL_PDF_Q] or '').strip(),
        'answer_pdf_url': str(row[COL_PDF_R] or '').strip(),
    }


def normalize(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Transform raw data into standard schema."""
    legislature = raw.get('legislature', '')
    question_number = raw.get('question_number', '')

    # Generate unique ID
    _id = f"LU-QP-{legislature}-{question_number}"

    # Get question and answer text
    question_text = raw.get('question_text', '') or ''
    answer_text = raw.get('answer_text', '') or ''

    # Build combined full text
    full_text_parts = []
    if question_text:
        full_text_parts.append(f"QUESTION:\n{question_text}")
    if answer_text:
        full_text_parts.append(f"\nRÉPONSE:\n{answer_text}")

    full_text = '\n'.join(full_text_parts)

    # Parse dates
    question_date = parse_date(raw.get('question_date'))
    answer_date = parse_date(raw.get('answer_date'))

    # Use question date as primary date
    date = question_date or answer_date

    return {
        '_id': _id,
        '_source': SOURCE_ID,
        '_type': 'doctrine',
        '_fetched_at': datetime.now(timezone.utc).isoformat(),
        'title': raw.get('title', '') or f"Question parlementaire {question_number}",
        'text': full_text,
        'date': date,
        'url': raw.get('question_url', ''),
        # Additional metadata
        'legislature': legislature,
        'question_number': question_number,
        'author': raw.get('author', ''),
        'political_party': raw.get('political_party', ''),
        'languages': raw.get('languages', ''),
        'question_date': question_date,
        'answer_date': answer_date,
        'answer_author': raw.get('answer_author', ''),
        'question_pdf_url': raw.get('question_pdf_url', ''),
        'answer_pdf_url': raw.get('answer_pdf_url', ''),
    }


def fetch_all(legislatures: list = None) -> Generator[Dict[str, Any], None, None]:
    """Fetch all parliamentary questions from specified legislatures."""
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'WorldWideLaw/1.0 (https://github.com/legal-data-hunter)',
        'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, */*',
    })

    if legislatures is None:
        legislatures = list(RESOURCES.keys())

    for legislature in legislatures:
        if legislature not in RESOURCES:
            print(f"  Warning: Unknown legislature {legislature}")
            continue

        url = RESOURCES[legislature]
        print(f"\nFetching legislature {legislature}...")

        try:
            xlsx_content = fetch_xlsx(url, session)
            wb = load_workbook(io.BytesIO(xlsx_content), data_only=True, read_only=True)
            ws = wb.active

            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header
                raw = parse_row(row)
                if raw:
                    yield raw
                    row_count += 1

            print(f"  Processed {row_count} records")
            wb.close()

        except Exception as e:
            print(f"  Error fetching {legislature}: {e}")
            import traceback
            traceback.print_exc()
            continue

        # Delay between legislatures
        time.sleep(1)


def fetch_updates(since: str) -> Generator[Dict[str, Any], None, None]:
    """Fetch documents modified since a given date.

    Note: The data.public.lu platform doesn't provide incremental updates,
    so this fetches all data from the most recent legislature and filters
    by question date.
    """
    since_date = datetime.fromisoformat(since.replace('Z', '+00:00'))

    # Fetch from most recent legislature only
    for raw in fetch_all(legislatures=['2023_2028']):
        question_date = parse_date(raw.get('question_date'))
        if question_date:
            try:
                doc_date = datetime.fromisoformat(question_date)
                if doc_date >= since_date.replace(tzinfo=None):
                    yield raw
            except:
                pass


def save_sample(record: Dict[str, Any], sample_dir: Path) -> None:
    """Save a normalized record to the sample directory."""
    filename = f"{record['_id']}.json"
    # Clean filename
    filename = re.sub(r'[^\w\-\.]', '_', filename)
    filepath = sample_dir / filename

    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(record, f, ensure_ascii=False, indent=2)


def bootstrap(sample_size: int = 15, legislature: str = None) -> None:
    """Bootstrap the data source with sample records."""
    script_dir = Path(__file__).parent
    sample_dir = script_dir / 'sample'
    sample_dir.mkdir(exist_ok=True)

    # Clear existing samples
    for f in sample_dir.glob('*.json'):
        f.unlink()

    print(f"\n=== Bootstrapping {SOURCE_ID} ===\n")
    print(f"Sample size: {sample_size}")

    legislatures = [legislature] if legislature else ['2023_2028']  # Most recent by default

    count = 0
    total_text_len = 0

    for raw in fetch_all(legislatures=legislatures):
        normalized = normalize(raw)

        # Check that we have meaningful text
        text_len = len(normalized.get('text', ''))
        if text_len < 100:  # Skip records with very short text
            continue

        save_sample(normalized, sample_dir)
        count += 1
        total_text_len += text_len

        print(f"  [{count}/{sample_size}] {normalized['_id']} - {text_len:,} chars")

        if count >= sample_size:
            break

    if count > 0:
        avg_text = total_text_len // count
        print(f"\n=== Complete ===")
        print(f"Records saved: {count}")
        print(f"Average text length: {avg_text:,} chars")
        print(f"Sample directory: {sample_dir}")
    else:
        print("\nNo records fetched!")


def main():
    parser = argparse.ArgumentParser(description='LU/Parliament data fetcher')
    parser.add_argument('command', choices=['bootstrap', 'fetch', 'updates'],
                        help='Command to execute')
    parser.add_argument('--sample', '-n', type=int, default=15,
                        help='Number of sample records (default: 15)')
    parser.add_argument('--legislature', '-l', type=str,
                        choices=['2013_2018', '2018_2023', '2023_2028'],
                        help='Specific legislature to fetch')
    parser.add_argument('--since', '-s', type=str,
                        help='Date for updates (ISO format)')

    args = parser.parse_args()

    if args.command == 'bootstrap':
        bootstrap(sample_size=args.sample, legislature=args.legislature)
    elif args.command == 'fetch':
        for raw in fetch_all(legislatures=[args.legislature] if args.legislature else None):
            normalized = normalize(raw)
            print(json.dumps(normalized, ensure_ascii=False))
    elif args.command == 'updates':
        if not args.since:
            print("Error: --since required for updates command")
            sys.exit(1)
        for raw in fetch_updates(args.since):
            normalized = normalize(raw)
            print(json.dumps(normalized, ensure_ascii=False))


if __name__ == '__main__':
    main()
